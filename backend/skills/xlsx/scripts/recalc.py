"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import contextlib
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path

from office.soffice import get_soffice_env, run_soffice

from openpyxl import load_workbook

MACRO_FILENAME = "Module1.xba"
SOFFICE_MISSING = "soffice not found on PATH; LibreOffice is required to recalculate"

MAX_LOCATIONS = 100

EXTERNAL_REF_RE = re.compile(r"""(?<![\w"\[])'?\[\d+\][^!"\[\]]*'?!""")

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _stamp(path):
    st = os.stat(path)
    return st.st_mtime_ns, st.st_size


def setup_libreoffice_macro(profile_dir: Path, timeout=30):
    url = profile_dir.as_uri()
    try:
        run_soffice(
            ["--headless", "--terminate_after_init", f"-env:UserInstallation={url}"],
            capture_output=True,
            timeout=timeout,
        )
    except FileNotFoundError:
        return None, SOFFICE_MISSING
    except subprocess.TimeoutExpired:
        return None, "LibreOffice timed out creating its profile; formulas were NOT recalculated"

    macro_dir = profile_dir / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        return None, "LibreOffice did not create a usable profile; formulas were NOT recalculated"

    try:
        (macro_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO)
    except OSError as e:
        return None, f"Could not install the recalculation macro: {e}"

    return url, None


def external_links_at_risk(filename):
    try:
        with zipfile.ZipFile(filename) as archive:
            names = archive.namelist()
    except (zipfile.BadZipFile, OSError):
        return []
    if not any(n.startswith("xl/externalLinks/") for n in names):
        return []

    with contextlib.ExitStack() as stack:
        formulas = load_workbook(filename, data_only=False)
        stack.callback(formulas.close)
        values = load_workbook(filename, data_only=True)
        stack.callback(values.close)

        external_names = [
            name
            for name, dn in formulas.defined_names.items()
            if isinstance(getattr(dn, "value", None), str) and EXTERNAL_REF_RE.search(dn.value)
        ]
        name_re = (
            re.compile(r"\b(" + "|".join(re.escape(n) for n in external_names) + r")\b")
            if external_names
            else None
        )

        at_risk = []
        for sheet in formulas.sheetnames:
            ws = formulas[sheet]
            if not hasattr(ws, "iter_rows"):  
                continue
            cached = values[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    reaches_out = EXTERNAL_REF_RE.search(v) or (name_re and name_re.search(v))
                    if reaches_out and cached[cell.coordinate].value is None:
                        at_risk.append(f"{sheet}!{cell.coordinate}")
        return at_risk


def recalc(filename, timeout=30, force=False):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not os.access(abs_path, os.W_OK):
        return {"error": f"{filename} is not writable; recalculation rewrites the file in place"}

    try:
        get_soffice_env()
    except Exception as e:  
        return {"error": f"Could not prepare the LibreOffice environment: {e}"}

    if not force:
        try:
            at_risk = external_links_at_risk(filename)
        except Exception as e:  
            return {"error": f"Could not inspect {filename} for external links: {e}"}
        if at_risk:
            shown = at_risk[:MAX_LOCATIONS]
            return {
                "error": (
                    "Refusing to recalculate: this workbook links to another workbook, and "
                    f"{len(at_risk)} linked cell(s) have lost their cached value (openpyxl strips "
                    "these on save). Recalculating would resolve them to #NAME? and delete the "
                    "external links for good. Copy those cells' values from the original file "
                    "before saving, or pass --force to accept the loss. Charts and conditional "
                    "formats can hold external references too, so this list may not be exhaustive."
                ),
                "external_link_cells": shown,
                "external_link_cells_truncated": max(0, len(at_risk) - len(shown)),
            }

    with tempfile.TemporaryDirectory(
        prefix="recalc-lo-profile-", ignore_cleanup_errors=True
    ) as profile_dir:
        return _recalc_with_profile(filename, abs_path, timeout, Path(profile_dir))


def _recalc_with_profile(filename, abs_path, timeout, profile_dir: Path):
    started = time.monotonic()
    profile_url, err = setup_libreoffice_macro(profile_dir, timeout=timeout)
    if err:
        return {"error": err}

    timeout = max(5, int(timeout - (time.monotonic() - started)))

    before = _stamp(abs_path)

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        f"-env:UserInstallation={profile_url}",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux" and shutil.which("timeout"):
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    timed_out = f"LibreOffice timed out after {timeout}s; formulas were NOT recalculated. Re-run with a longer timeout."

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, env=get_soffice_env(), timeout=timeout + 15
        )
    except subprocess.TimeoutExpired:
        return {"error": timed_out}
    except FileNotFoundError:
        return {"error": SOFFICE_MISSING}

    if result.returncode == 124:
        return {"error": timed_out}

    if result.returncode != 0:
        detail = (result.stderr or "").strip() or f"soffice exited {result.returncode}"
        return {"error": f"LibreOffice failed to recalculate: {detail}"}

    if _stamp(abs_path) == before:
        return {
            "error": (
                "LibreOffice exited cleanly but never rewrote the file, so nothing was "
                "recalculated. Check that no other LibreOffice instance is running, then retry."
            )
        }

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):  
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                entry = {"count": len(locations), "locations": locations[:MAX_LOCATIONS]}
                if len(locations) > MAX_LOCATIONS:
                    entry["locations_truncated"] = len(locations) - MAX_LOCATIONS
                result["error_summary"][err_type] = entry

        wb.close()

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            if not hasattr(ws, "iter_rows"):  
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    if not args:
        print("Usage: python recalc.py <excel_file> [timeout_seconds] [--force]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        print("\nOn any failure the JSON has an 'error' key and no 'status'.")
        print("--force recalculates even when it would destroy external links.")
        sys.exit(1)

    filename = args[0]
    timeout = int(args[1]) if len(args) > 1 else 30

    result = recalc(filename, timeout, force=force)
    print(json.dumps(result, indent=2))
    sys.exit(1 if "error" in result else 0)


if __name__ == "__main__":
    main()
